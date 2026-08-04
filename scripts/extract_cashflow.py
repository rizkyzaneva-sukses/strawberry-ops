"""
Ekstraksi "CASHFLOW The Red Harvest.xlsx" menjadi JSON ternormalisasi.

Sheet sumber punya tiga format tabel gaji yang berbeda plus blok rekapitulasi
harian, jadi tiap region diparsing terpisah lalu digabung.

Pakai:
    python scripts/extract_cashflow.py "C:/path/CASHFLOW The Red Harvest.xlsx"

Hasil ditulis ke prisma/import-data/*.json
"""

import json
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "prisma" / "import-data"

KEBUN_BARU = "KEBUN_BARU"
KEBUN_LAMA = "KEBUN_LAMA"
SEMUA = "SEMUA"

warnings = []


def warn(kind, message, **extra):
    entry = {"kind": kind, "message": message}
    entry.update(extra)
    warnings.append(entry)


# ---------------------------------------------------------------- utilities


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        text = unicodedata.normalize("NFKC", value).replace("\xa0", " ").strip()
        return text or None
    return value


def as_int(value):
    if value is None or isinstance(value, str):
        return None
    return int(round(float(value)))


def as_float(value):
    if value is None or isinstance(value, str):
        return None
    return float(value)


DATE_RE = re.compile(r"(\d{1,2})\s*/\s*(\d{1,2})\s*/\s*(\d{4})")


def parse_date(value, ref=None):
    """Tanggal di sheet ditulis sebagai teks dd/mm/yyyy, tapi sebagian
    terlanjur dikonversi Excel memakai urutan m/d/yyyy sehingga hari dan
    bulannya tertukar. Konversi datetime selalu ditukar balik -- Excel hanya
    salah baca ketika hari <= 12, dan semua kasus di sheet ini memang begitu.
    """
    if value is None:
        return None

    if isinstance(value, (datetime, date)):
        d = value.date() if isinstance(value, datetime) else value
        if d.day <= 12:
            fixed = date(d.year, d.day, d.month)
            warn(
                "tanggal-ditukar",
                f"Excel membaca {d.isoformat()} sebagai m/d; dikembalikan ke {fixed.isoformat()}",
                ref=ref,
                asli=d.isoformat(),
                hasil=fixed.isoformat(),
            )
            return fixed
        warn(
            "tanggal-ambigu",
            f"Tanggal {d.isoformat()} tidak bisa ditukar (hari > 12), dipakai apa adanya",
            ref=ref,
        )
        return d

    text = clean(value)
    if not text:
        return None
    match = DATE_RE.search(text)
    if not match:
        warn("tanggal-gagal", f"Tidak bisa membaca tanggal dari {text!r}", ref=ref)
        return None
    day, month, year = (int(g) for g in match.groups())
    try:
        return date(year, month, day)
    except ValueError:
        warn("tanggal-gagal", f"Tanggal tidak valid: {text!r}", ref=ref)
        return None


def iso(d):
    return d.isoformat() if d else None


def norm_garden(value):
    text = (clean(value) or "").upper()
    if "BARU" in text:
        return KEBUN_BARU
    if "LAMA" in text:
        return KEBUN_LAMA
    if "SEMUA" in text:
        return SEMUA
    return None


TIME_RE = re.compile(r"(\d{1,2})[.:](\d{2})")


def parse_time_range(value):
    """'07.00 - 12.00' -> (7.0, 12.0). Mengembalikan None kalau borongan."""
    text = clean(value)
    if not isinstance(text, str) or not text:
        return None
    hits = TIME_RE.findall(text)
    if len(hits) < 2:
        return None
    start = int(hits[0][0]) + int(hits[0][1]) / 60
    end = int(hits[1][0]) + int(hits[1][1]) / 60
    if end < start:
        end += 12
    return start, end


NGABEDUG_END = 12.0
NYORE_END = 15.0


def classify_shift(span):
    """Kembalikan (shift, lembur_hours) dari rentang waktu."""
    if span is None:
        return "BORONGAN", 0.0
    start, end = span
    if start < NGABEDUG_END:
        return "NGABEDUG", 0.0
    if start < NYORE_END:
        return "NYORE", max(0.0, round(end - NYORE_END, 2))
    return "LEMBUR", max(0.0, round(end - start, 2))


# ------------------------------------------------------------ master pekerja

GROUP_EMPLOYEE = "Pekerja Harian (Grup)"

GROUP_RE = re.compile(r"^(perempuan|laki2?\.?|laki-laki)\s*[x×]\s*(\d+)", re.IGNORECASE)

# Baris rekapitulasi yang bentuknya mirip baris pekerja
JUNK_NAME_RE = re.compile(
    r"^(total|nama pekerja|periode gaji|sudah di tf|bensin|dana talang|cash ke|dll)\b",
    re.IGNORECASE,
)


def is_junk_name(name):
    return bool(name) and bool(JUNK_NAME_RE.match(name))


def read_employees(ws):
    employees = {}
    for row in range(7, 20):
        name = clean(ws.cell(row, 4).value)
        if not name:
            continue
        monthly = as_int(ws.cell(row, 9).value) or 0
        employees[name] = {
            "fullName": name,
            "defaultJobs": clean(ws.cell(row, 5).value),
            "wageNgabedug": as_int(ws.cell(row, 6).value) or 0,
            "wageNyore": as_int(ws.cell(row, 7).value) or 0,
            "wageLemburPerHour": as_int(ws.cell(row, 8).value) or 0,
            "monthlySalary": monthly,
            "employmentType": "BULANAN" if monthly else "HARIAN",
            "gender": "P" if name.lower().startswith("bu ") else "L",
            "isGroup": False,
        }
    return employees


def ensure_employee(employees, name):
    """Pekerja yang muncul di log tapi tidak ada di master ditambahkan
    dengan tarif standar rekan sejenisnya."""
    if name in employees:
        return employees[name]
    female = name.lower().startswith("bu ")
    employees[name] = {
        "fullName": name,
        "defaultJobs": None,
        "wageNgabedug": 45000 if female else 50000,
        "wageNyore": 30000 if female else 20000,
        "wageLemburPerHour": 0 if female else 10000,
        "monthlySalary": 0,
        "employmentType": "HARIAN",
        "gender": "P" if female else "L",
        "isGroup": False,
    }
    warn(
        "pekerja-baru",
        f"{name!r} ada di log kerja tapi tidak ada di master DATA PEKERJA; "
        "ditambahkan dengan tarif standar - mohon dicek",
    )
    return employees[name]


def resolve_worker(raw, employees):
    """Kembalikan (nama_karyawan, headcount). Entri kolektif dipetakan ke
    karyawan grup."""
    text = clean(raw)
    if not text:
        return None, 1

    group = GROUP_RE.match(text)
    if group:
        return GROUP_EMPLOYEE, int(group.group(2))

    low = text.lower()
    if low.startswith("laki2. ") or low.startswith("laki2 "):
        text = text.split(None, 1)[1]

    for known in employees:
        if known.lower() == text.lower():
            return known, 1
    for known in employees:
        if known.lower() in text.lower():
            return known, 1
    return text, 1


# ----------------------------------------------------------- sheet Arus Kas


def read_budget_items(ws):
    """Blok DATA UTAMA (INPUT): anggaran capex + harga faktual."""
    items = []
    for row in range(4, 10):
        name = clean(ws.cell(row, 3).value)
        if not name:
            continue
        qty = as_float(ws.cell(row, 4).value)
        if qty is None:
            raw = clean(ws.cell(row, 4).value)
            qty = float(re.sub(r"[^\d.]", "", raw)) if raw else 0.0
        planned_price = as_int(ws.cell(row, 5).value) or 0
        planned_total = as_int(ws.cell(row, 6).value) or 0
        status = clean(ws.cell(row, 7).value)
        actual_price = as_int(ws.cell(row, 8).value)
        variance = as_int(ws.cell(row, 9).value)
        items.append(
            {
                "garden": KEBUN_BARU,
                "name": name,
                "plannedQty": qty,
                "plannedUnitPrice": planned_price,
                "plannedTotal": planned_total,
                "actualUnitPrice": actual_price,
                "actualTotal": int(actual_price * qty) if actual_price else None,
                "variance": variance,
                "paymentStatus": "LUNAS" if status == "LUNAS" else "BELUM_BAYAR",
                "sortOrder": row,
            }
        )
    return items


def read_assets(ws):
    """Blok ALAT-ALAT. Kolom Jml berisi porsi kepemilikan (0.5 = patungan)."""
    assets = []
    for row in range(13, 16):
        name = clean(ws.cell(row, 3).value)
        if not name:
            continue
        share = as_float(ws.cell(row, 4).value) or 1.0
        unit_price = as_int(ws.cell(row, 5).value) or 0
        total = as_int(ws.cell(row, 6).value) or 0
        status = clean(ws.cell(row, 7).value)
        assets.append(
            {
                "garden": KEBUN_BARU if share >= 1 else None,
                "name": name,
                "category": "MESIN",
                "quantity": 1,
                "unitPrice": unit_price,
                "ownershipShare": share,
                "totalCost": total,
                "paymentStatus": "LUNAS" if status == "LUNAS" else "BELUM_BAYAR",
                "notes": None
                if share >= 1
                else f"Porsi kepemilikan {share:g} - patungan dengan kebun lain",
            }
        )
    return assets


def read_capital(ws):
    """Blok DANA MASUK INVESTOR. Kolom G = modal penyertaan,
    kolom H = modal kasbon (utang)."""
    entries = []
    for row in range(19, 28):
        when = parse_date(ws.cell(row, 3).value, ref=f"Arus Kas!C{row}")
        description = clean(ws.cell(row, 4).value)
        if not description:
            continue
        source = clean(ws.cell(row, 5).value)
        destination = clean(ws.cell(row, 6).value)
        proof = clean(ws.cell(row, 10).value)
        note = clean(ws.cell(row, 11).value)
        for column, funding in ((7, "EQUITY"), (8, "LOAN")):
            amount = as_int(ws.cell(row, column).value)
            if not amount:
                continue
            entries.append(
                {
                    "garden": KEBUN_BARU,
                    "entryDate": iso(when),
                    "description": description,
                    "fundingType": funding,
                    "amount": amount,
                    "sourceAccount": source,
                    "destinationAccount": destination,
                    "proofRef": proof,
                    "notes": note,
                }
            )
    return entries


FLAG_HINTS = ("penipu", "tidak ada bukti", "tidak memberikan nota", "fiktif", "markup", "downgrade")


def read_operational_expenses(ws):
    """Blok DANA KELUAR OPERASIONAL.

    Baris baru dianggap transaksi tersendiri bila kolom Sumber Dana terisi;
    kalau tidak, baris itu rincian tambahan dari transaksi sebelumnya
    (satu transfer untuk beberapa keperluan).
    """
    transactions = []
    current = None
    current_date = None

    for row in range(32, 105):
        description = clean(ws.cell(row, 4).value)
        source = clean(ws.cell(row, 5).value)
        amount = as_int(ws.cell(row, 7).value)
        if not description:
            continue
        if description.startswith("Total"):
            continue

        when = parse_date(ws.cell(row, 3).value, ref=f"Arus Kas!C{row}")
        if when:
            current_date = when
        else:
            # Sebagian baris transaksi tidak mengulang tanggalnya
            when = current_date
        destination = clean(ws.cell(row, 6).value)
        proof = clean(ws.cell(row, 10).value)
        note = clean(ws.cell(row, 11).value)

        if source:
            current = {
                "garden": KEBUN_BARU,
                "transactionDate": iso(when),
                "description": description,
                "amount": amount or 0,
                "sourceAccount": source,
                "recipientAccount": destination,
                "proofRef": proof,
                "notes": note,
                "isFlagged": bool(note and any(h in note.lower() for h in FLAG_HINTS)),
                "flagNote": note if note and any(h in note.lower() for h in FLAG_HINTS) else None,
                "row": row,
                "items": [],
            }
            transactions.append(current)
        elif current is not None:
            current["items"].append(
                {
                    "description": description,
                    "amount": amount or 0,
                    "proofRef": proof,
                }
            )
            if amount:
                current["amount"] += amount
            if note and not current["notes"]:
                current["notes"] = note
        else:
            warn("pengeluaran-yatim", f"Baris {row} tanpa induk transaksi", ref=f"Arus Kas!D{row}")

    for txn in transactions:
        if not txn["items"]:
            txn.pop("items")
        txn.pop("row")
    return transactions


PAYMENT_STATUS_RE = [
    (re.compile(r"\bLUNAS\b", re.IGNORECASE), "LUNAS"),
    (re.compile(r"\bDP\b", re.IGNORECASE), "DP"),
    (re.compile(r"kurang bayar", re.IGNORECASE), "KURANG_BAYAR"),
]

STAGE_RE = re.compile(r"(Tahap\s*\d+|Pelunasan|DP)", re.IGNORECASE)

CATEGORY_RULES = [
    ("LAHAN", ("lahan",)),
    ("BIBIT", ("bibit",)),
    ("JASA_NGARUNG", ("ngarung",)),
    ("KARUNG", ("karung",)),
    ("SAUNG", ("saung", "triplek", "karpet", "vinyl")),
    ("ALAT", ("mesin", "pompa", "sprayer", "terpal", "selang", "water")),
    ("PUPUK_OBAT", ("pupuk", "obat")),
    ("UPAH_HARIAN", ("harian", "tanam", "siram", "babad", "penampungan air")),
    ("BENSIN", ("bensin",)),
]

# Awal periode rincian harian dicatat di sheet Operasional
DETAIL_START = "2026-06-01"


def classify_category(text):
    low = (text or "").lower()
    for code, keywords in CATEGORY_RULES:
        if any(keyword in low for keyword in keywords):
            return code
    return "LAIN_LAIN"


def enrich_expense(txn):
    text = txn["description"] or ""
    notes = txn.get("notes") or ""
    txn["paymentStatus"] = "LUNAS"
    for pattern, status in PAYMENT_STATUS_RE:
        if pattern.search(text):
            txn["paymentStatus"] = status
            break
    stage = STAGE_RE.search(text)
    txn["installmentLabel"] = stage.group(1) if stage else None

    # Transfer "Operasional Kebun N" yang isinya penggantian gaji harian sudah
    # terwakili oleh PayrollRecord. Ditandai supaya laporan biaya tidak
    # menghitungnya dua kali, tapi tetap muncul di laporan arus kas.
    reimbursement = any(word in notes.lower() for word in ("gaji", "harian"))
    in_detail_period = (txn.get("transactionDate") or "") >= DETAIL_START
    txn["isAdvance"] = bool(
        reimbursement
        and in_detail_period
        and text.lower().startswith("operasional kebun")
    )
    txn["category"] = "OPERASIONAL_HARIAN" if txn["isAdvance"] else classify_category(
        f"{text} {notes}"
    )
    return txn


# --------------------------------------------------------- sheet Operasional


def read_harvests(ws):
    """Blok DATA TRANSAKSIONAL BANDAR (kolom R-AJ)."""
    harvests = []
    for row in range(9, ws.max_row + 1):
        garden = norm_garden(ws.cell(row, 20).value)
        if garden not in (KEBUN_BARU, KEBUN_LAMA):
            continue
        when = parse_date(ws.cell(row, 19).value, ref=f"Operasional!S{row}")
        if not when:
            continue

        lama = garden == KEBUN_LAMA
        normal_kg = as_float(ws.cell(row, 21 if lama else 22).value) or 0.0
        bs_kg = as_float(ws.cell(row, 26 if lama else 27).value) or 0.0
        normal_price = as_int(ws.cell(row, 23).value) or 0
        bs_price = as_int(ws.cell(row, 28).value) or 0
        normal_revenue = as_int(ws.cell(row, 24 if lama else 25).value) or 0
        bs_revenue = as_int(ws.cell(row, 29 if lama else 30).value) or 0
        total_kg = as_float(ws.cell(row, 31 if lama else 32).value)
        total_revenue = as_int(ws.cell(row, 33 if lama else 34).value)

        if total_kg is None:
            total_kg = round(normal_kg + bs_kg, 2)
        if total_revenue is None:
            total_revenue = normal_revenue + bs_revenue

        harvests.append(
            {
                "garden": garden,
                "harvestDate": iso(when),
                "normalPricePerKg": normal_price,
                "bsPricePerKg": bs_price,
                "totalHarvestKg": round(total_kg, 2),
                "bsKg": round(bs_kg, 2),
                "normalKg": round(normal_kg, 2),
                "normalRevenue": normal_revenue,
                "bsRevenue": bs_revenue,
                "totalRevenue": total_revenue,
                "bsPercentage": round(bs_kg / total_kg, 6) if total_kg else 0.0,
            }
        )
    return harvests


def make_payroll(employee, when, garden, job, span, wage, headcount=1, notes=None, manual=False):
    shift, lembur = classify_shift(span)
    return {
        "employee": employee,
        "workDate": iso(when),
        "garden": garden,
        "jobType": job,
        "shift": shift,
        "startTime": f"{int(span[0]):02d}.{int(round((span[0] % 1) * 60)):02d}" if span else None,
        "endTime": f"{int(span[1]):02d}.{int(round((span[1] % 1) * 60)):02d}" if span else None,
        "lemburHours": lembur,
        "headcount": headcount,
        "wageAmount": int(wage),
        "isManualWage": manual,
        "notes": notes,
    }


def expected_wage(profile, span):
    shift, lembur = classify_shift(span)
    if shift == "NGABEDUG":
        return profile["wageNgabedug"]
    if shift == "NYORE":
        return profile["wageNyore"] + int(lembur * profile["wageLemburPerHour"])
    if shift == "LEMBUR":
        return int(lembur * profile["wageLemburPerHour"])
    return 0


def read_payroll_table_flat(ws, header_row, end_row, employees, fuel, salary_columns):
    """Format tabel awal (satu baris = satu entri lengkap).

    Kolom: C No, D Tgl, E Nama, F Pekerjaan, G Area, H Blok, I Waktu,
    salary_columns, M Cost Kb.baru, N Cost Kb.lama.
    """
    records = []
    current_date = None

    for row in range(header_row + 1, end_row + 1):
        when = parse_date(ws.cell(row, 4).value, ref=f"Operasional!D{row}")
        if when:
            current_date = when
        raw_name = clean(ws.cell(row, 5).value)
        if not raw_name or current_date is None:
            continue

        salary = None
        for column in salary_columns:
            value = as_int(ws.cell(row, column).value)
            if value:
                salary = value
                break
        if not salary:
            continue

        if "bensin" in raw_name.lower():
            fuel.append(build_fuel(current_date, raw_name, salary, ws.cell(row, 7).value,
                                   as_int(ws.cell(row, 13).value), as_int(ws.cell(row, 14).value)))
            continue

        if is_junk_name(raw_name):
            continue

        employee, headcount = resolve_worker(raw_name, employees)
        ensure_employee(employees, employee)
        job = clean(ws.cell(row, 6).value)
        garden = norm_garden(ws.cell(row, 7).value)
        span = parse_time_range(ws.cell(row, 9).value)
        baru = as_int(ws.cell(row, 13).value) or 0
        lama = as_int(ws.cell(row, 14).value) or 0

        if garden == SEMUA and (baru or lama):
            # Sheet sudah memecah biayanya; ikuti pembagian itu.
            for target, amount in ((KEBUN_BARU, baru), (KEBUN_LAMA, lama)):
                if amount:
                    records.append(
                        make_payroll(employee, current_date, target, job, span, amount, headcount)
                    )
            continue

        records.append(
            make_payroll(
                employee,
                current_date,
                garden or KEBUN_BARU,
                job,
                span,
                salary,
                headcount,
            )
        )
    return records


def build_fuel(when, label, amount, area, baru, lama):
    litre_match = re.search(r"(\d+(?:[.,]\d+)?)\s*ltr", label or "", re.IGNORECASE)
    litres = float(litre_match.group(1).replace(",", ".")) if litre_match else None
    garden = norm_garden(area) or SEMUA
    return {
        "transactionDate": iso(when),
        "category": "BENSIN",
        "description": label or "Bensin",
        "amount": int(amount),
        "quantity": litres,
        "unit": "Liter" if litres else None,
        "unitPrice": int(round(amount / litres)) if litres else None,
        "garden": garden,
        "allocations": {KEBUN_BARU: baru or 0, KEBUN_LAMA: lama or 0}
        if garden == SEMUA and (baru or lama)
        else None,
    }


def read_payroll_table_blocks(ws, header_row, end_row, employees, fuel, salary_columns):
    """Format tabel dengan blok per orang.

    Nama ada di baris pertama blok, baris berikutnya adalah shift lanjutan.
    Untuk pekerja perempuan nominalnya hanya ditulis sekali di baris terakhir
    sebagai total harian, sedangkan untuk pekerja laki-laki ditulis per baris.
    """
    records = []
    current_date = None
    block = []
    block_name = None

    def flush():
        nonlocal block, block_name
        if block_name and block and not is_junk_name(block_name):
            records.extend(
                resolve_block(block_name, block, current_date, employees, fuel)
            )
        block = []
        block_name = None

    for row in range(header_row + 1, end_row + 1):
        when = parse_date(ws.cell(row, 4).value, ref=f"Operasional!D{row}")
        raw_name = clean(ws.cell(row, 5).value)

        # Baris subtotal mingguan mengisi beberapa kolom upah sekaligus,
        # baris pekerja tidak pernah begitu.
        filled = [column for column in salary_columns if as_int(ws.cell(row, column).value)]
        if len(filled) > 1:
            continue

        if raw_name:
            flush()
            block_name = raw_name
        if when:
            current_date = when

        job = clean(ws.cell(row, 6).value)
        garden = norm_garden(ws.cell(row, 7).value)
        span_raw = ws.cell(row, 9).value
        salary = as_int(ws.cell(row, filled[0]).value) if filled else None

        if job or span_raw or salary is not None:
            block.append(
                {
                    "row": row,
                    "job": job,
                    "garden": garden,
                    "span_raw": span_raw,
                    "span": parse_time_range(span_raw),
                    "salary": salary,
                    "baru": as_int(ws.cell(row, 13).value),
                    "lama": as_int(ws.cell(row, 14).value),
                }
            )

    flush()
    return records


def resolve_block(raw_name, lines, when, employees, fuel):
    if when is None:
        return []

    # Baris bensin di dalam blok orang dicatat sebagai pengeluaran, bukan upah.
    kept = []
    for line in lines:
        label = clean(line["span_raw"])
        label = label if isinstance(label, str) else ""
        job_label = line["job"] if isinstance(line["job"], str) else ""
        if "bensin" in label.lower() or "bensin" in job_label.lower():
            amount = line["salary"] or (line["baru"] or 0) + (line["lama"] or 0)
            if amount:
                fuel.append(
                    build_fuel(
                        when, label or job_label, amount, None, line["baru"], line["lama"]
                    )
                )
            continue
        kept.append(line)

    work = [line for line in kept if line["span"] or line["job"]]
    if not work:
        return []

    employee, headcount = resolve_worker(raw_name, employees)
    profile = ensure_employee(employees, employee)

    # Upah bisa ditulis per baris (pekerja laki-laki) atau sekali saja sebagai
    # total harian (pekerja perempuan, dan blok yang nominalnya ditaruh di
    # baris kosong setelah daftar pekerjaan).
    salaried_work = [line for line in work if line["salary"]]
    trailing = [line for line in kept if line["salary"] and line not in work]
    shift_groups = {(classify_shift(line["span"])[0], line["garden"]) for line in work}

    if trailing and not salaried_work:
        return distribute_block(
            employee, profile, when, work, sum(l["salary"] for l in trailing), headcount
        )
    if len(salaried_work) == 1 and len(shift_groups) > 1:
        return distribute_block(
            employee, profile, when, work, salaried_work[0]["salary"], headcount
        )

    records = []
    for line in work:
        if not line["salary"]:
            continue
        garden = line["garden"] or KEBUN_BARU
        if garden == SEMUA and (line["baru"] or line["lama"]):
            for target, amount in ((KEBUN_BARU, line["baru"]), (KEBUN_LAMA, line["lama"])):
                if amount:
                    records.append(
                        make_payroll(employee, when, target, line["job"], line["span"], amount, headcount)
                    )
            continue
        records.append(
            make_payroll(employee, when, garden, line["job"], line["span"], line["salary"], headcount)
        )
    return records


def duration(line):
    span = line["span"]
    return (span[1] - span[0]) if span else 1.0


def distribute_block(employee, profile, when, lines, total, headcount):
    """Satu nominal harian dipecah ke tiap shift memakai tarif karyawan.

    Satu shift dibayar satu tarif berapa pun jumlah pekerjaannya, jadi
    pengelompokan dilakukan per shift lebih dulu. Kalau dalam satu shift
    pekerjanya berpindah kebun, jatah shift itu dibagi menurut lama kerja.
    """
    by_shift = defaultdict(list)
    for line in lines:
        by_shift[classify_shift(line["span"])[0]].append(line)

    weights = {}
    derived = True
    for shift, members in by_shift.items():
        rate = max(expected_wage(profile, member["span"]) for member in members)
        if not rate:
            derived = False
        weights[shift] = rate or 1

    weight_sum = sum(weights.values())
    records = []
    remaining = total

    for index, (shift, members) in enumerate(by_shift.items()):
        last = index == len(by_shift) - 1
        share = remaining if last else int(round(total * weights[shift] / weight_sum))
        remaining -= share

        by_garden = defaultdict(list)
        for member in members:
            by_garden[member["garden"] or KEBUN_BARU].append(member)

        hours_total = sum(duration(member) for member in members) or 1.0
        left = share
        for position, (garden, garden_lines) in enumerate(by_garden.items()):
            final = position == len(by_garden) - 1
            hours = sum(duration(line) for line in garden_lines)
            amount = left if final else int(round(share * hours / hours_total))
            left -= amount

            garden_lines.sort(key=duration, reverse=True)
            primary = garden_lines[0]
            extra = [line["job"] for line in garden_lines[1:] if line["job"]]
            note = ("Juga: " + ", ".join(extra)) if extra else None

            records.append(
                make_payroll(
                    employee, when, garden, primary["job"], primary["span"],
                    amount, headcount, note,
                )
            )

    expected_total = sum(weights.values())
    if derived and expected_total and abs(expected_total - total) > 1000:
        warn(
            "upah-tidak-cocok",
            f"{employee} {iso(when)}: total di sheet Rp{total:,} tapi tarif master "
            f"menghasilkan Rp{expected_total:,}",
            tanggal=iso(when),
            pekerja=employee,
        )
    return records


def split_shared_payroll(records):
    """Catatan yang areanya masih 'SEMUA KEBUN' dibagi rata dua kebun.

    Sisa pembulatan diberikan ke kebun baru supaya totalnya tetap utuh.
    """
    result = []
    shared = 0
    for record in records:
        if record["garden"] != SEMUA:
            result.append(record)
            continue
        shared += 1
        half = record["wageAmount"] // 2
        for garden, amount in ((KEBUN_LAMA, half), (KEBUN_BARU, record["wageAmount"] - half)):
            clone = dict(record)
            clone["garden"] = garden
            clone["wageAmount"] = amount
            clone["notes"] = " | ".join(
                filter(None, [record.get("notes"), "Dibagi rata dari kegiatan SEMUA KEBUN"])
            )
            result.append(clone)
    if shared:
        warn(
            "upah-dibagi-rata",
            f"{shared} catatan upah berarea SEMUA KEBUN dibagi 50/50 antara dua kebun",
        )
    return result


def read_daily_recap(ws, start_row, employees):
    """Blok REKAPITULASI PENGELUARAN KEBUN - satu blok per tanggal.

    Kolom: C Nama, D Pekerjaan, E Area, F Blok, G Waktu, H Pengeluaran,
    I Total, J Kb.baru, K Kb.lama.
    """
    records = []
    expenses = []
    periods = []

    current_date = None
    mode = None
    current_name = None

    row = start_row
    while row <= ws.max_row:
        label = clean(ws.cell(row, 3).value)
        if not isinstance(label, str):
            label = None

        if label == "Tgl":
            current_date = parse_date(ws.cell(row, 4).value, ref=f"Operasional!D{row}")
            mode = None
            current_name = None
        elif label == "GAJI KARYAWAN":
            mode = "payroll"
            current_name = None
            row += 1  # lewati baris header
        elif label == "PENGELUARAN LAIN-LAIN":
            mode = "expense"
            row += 1
        elif label == "PERIODE GAJI":
            period = read_period(ws, row)
            if period:
                periods.append(period)
            mode = None
        elif label and label.startswith("Total"):
            pass
        elif mode == "payroll" and current_date:
            if label:
                current_name = None if is_junk_name(label) else label
            amount = as_int(ws.cell(row, 8).value)
            if amount and current_name:
                employee, headcount = resolve_worker(current_name, employees)
                ensure_employee(employees, employee)
                garden = norm_garden(ws.cell(row, 5).value)
                baru = as_int(ws.cell(row, 10).value) or 0
                lama = as_int(ws.cell(row, 11).value) or 0
                job = clean(ws.cell(row, 4).value)
                span = parse_time_range(ws.cell(row, 7).value)
                block_name = clean(ws.cell(row, 6).value)

                if garden == SEMUA and (baru or lama):
                    for target, value in ((KEBUN_BARU, baru), (KEBUN_LAMA, lama)):
                        if value:
                            entry = make_payroll(
                                employee, current_date, target, job, span, value, headcount
                            )
                            entry["block"] = block_name
                            records.append(entry)
                else:
                    entry = make_payroll(
                        employee,
                        current_date,
                        garden or (KEBUN_BARU if baru else KEBUN_LAMA),
                        job,
                        span,
                        amount,
                        headcount,
                    )
                    entry["block"] = block_name
                    records.append(entry)
        elif mode == "expense" and current_date and label:
            total = as_int(ws.cell(row, 9).value) or as_int(ws.cell(row, 8).value)
            if total:
                garden = norm_garden(ws.cell(row, 5).value) or SEMUA
                baru = as_int(ws.cell(row, 10).value) or 0
                lama = as_int(ws.cell(row, 11).value) or 0
                quantity = as_float(ws.cell(row, 6).value)
                unit = clean(ws.cell(row, 7).value)
                expenses.append(
                    {
                        "transactionDate": iso(current_date),
                        "category": "BENSIN" if "bensin" in label.lower() else "LAIN_LAIN",
                        "description": label,
                        "amount": total,
                        "quantity": quantity,
                        "unit": unit,
                        "unitPrice": as_int(ws.cell(row, 8).value),
                        "garden": garden,
                        "allocations": {KEBUN_BARU: baru, KEBUN_LAMA: lama}
                        if garden == SEMUA and (baru or lama)
                        else None,
                    }
                )
        row += 1

    return records, expenses, periods


PERIOD_RE = re.compile(r"(\d{1,2}/\d{1,2}/\d{4}).*?(\d{1,2}/\d{1,2}/\d{4})")


def read_period(ws, row, name_col=3, value_col=4):
    text = clean(ws.cell(row + 1, name_col).value)
    if not text:
        return None
    match = PERIOD_RE.search(text)
    if not match:
        return None
    start = parse_date(match.group(1), ref=f"Operasional!C{row + 1}")
    end = parse_date(match.group(2), ref=f"Operasional!C{row + 1}")
    if not start or not end:
        return None

    lines = []
    payments = []
    cursor = row + 3
    while cursor <= ws.max_row:
        label = clean(ws.cell(cursor, name_col).value)
        if not isinstance(label, str):
            label = None
        amount = as_int(ws.cell(cursor, value_col).value)
        if label in (None, "") and amount is None:
            cursor += 1
            if cursor > row + 60:
                break
            continue
        if label and label.startswith("Tgl"):
            break
        if label and label.lower().startswith("sudah di tf"):
            batch = re.search(r"batch\s*(\d+)", label, re.IGNORECASE)
            payments.append(
                {
                    "batchNo": int(batch.group(1)) if batch else len(payments) + 1,
                    "amount": amount or 0,
                    "label": label,
                }
            )
        elif label and not label.startswith("Total") and label != "NAMA PEKERJA" and amount:
            lines.append({"employee": label, "amount": amount})
        cursor += 1

    return {
        "startDate": iso(start),
        "endDate": iso(end),
        "lines": lines,
        "payments": payments,
    }


def read_fuel_history(ws, start_row):
    """Blok HISTORI PEMBELIAN BENSIN."""
    entries = []
    for row in range(start_row + 2, start_row + 40):
        when = parse_date(ws.cell(row, 4).value, ref=f"Operasional!D{row}")
        litres = as_float(ws.cell(row, 5).value)
        price = as_int(ws.cell(row, 6).value)
        total = as_int(ws.cell(row, 7).value)
        if not when or not total:
            if clean(ws.cell(row, 3).value) is None and total is None:
                break
            continue
        entries.append(
            {
                "transactionDate": iso(when),
                "category": "BENSIN",
                "description": f"Bensin {litres:g} liter" if litres else "Bensin",
                "amount": total,
                "quantity": litres,
                "unit": "Liter",
                "unitPrice": price,
                "garden": SEMUA,
                "allocations": None,
            }
        )
    return entries


def read_side_periods(ws):
    """Blok PERIODE GAJI yang diletakkan di kolom E (region tabel awal)."""
    periods = []
    for row in range(1, ws.max_row + 1):
        if clean(ws.cell(row, 5).value) == "PERIODE GAJI":
            period = read_period(ws, row, name_col=5, value_col=6)
            if period:
                periods.append(period)
    return periods


# ---------------------------------------------------------------------- main


def main():
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
        r"C:\Users\rizky\Downloads\CASHFLOW The Red Harvest.xlsx"
    )
    workbook = openpyxl.load_workbook(source, data_only=True)
    kas = workbook["Arus Kas"]
    ops = workbook["Operasional"]

    employees = read_employees(ops)
    employees[GROUP_EMPLOYEE] = {
        "fullName": GROUP_EMPLOYEE,
        "defaultJobs": None,
        "wageNgabedug": 45000,
        "wageNyore": 30000,
        "wageLemburPerHour": 0,
        "monthlySalary": 0,
        "employmentType": "HARIAN",
        "gender": None,
        "isGroup": True,
    }

    fuel = []
    payroll = []

    # Tiga region dengan tata letak berbeda; batas bawah dipilih tepat sebelum
    # blok rekapitulasi mingguan yang menyusul.
    payroll += read_payroll_table_flat(ops, 38, 57, employees, fuel, [10])
    payroll += read_payroll_table_blocks(ops, 59, 589, employees, fuel, [10, 11])
    payroll += read_payroll_table_blocks(ops, 623, 806, employees, fuel, [10, 11, 12])

    recap_payroll, recap_expenses, periods = read_daily_recap(ops, 852, employees)
    payroll += recap_payroll

    payroll = split_shared_payroll(payroll)

    fuel += read_fuel_history(ops, 808)
    periods += read_side_periods(ops)

    capital = read_capital(kas)
    expenses = [enrich_expense(txn) for txn in read_operational_expenses(kas)]

    payload = {
        "employees.json": sorted(employees.values(), key=lambda e: e["fullName"]),
        "jobTypes.json": [
            clean(ops.cell(row, 4).value)
            for row in range(23, 38)
            if clean(ops.cell(row, 4).value)
        ],
        "budgetItems.json": read_budget_items(kas),
        "assets.json": read_assets(kas),
        "capitalInjections.json": capital,
        "operationalExpenses.json": expenses,
        "dailyExpenses.json": recap_expenses + fuel,
        "harvests.json": read_harvests(ops),
        "payrollRecords.json": payroll,
        "payrollPeriods.json": periods,
        "warnings.json": warnings,
    }

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for filename, data in payload.items():
        (OUT_DIR / filename).write_text(
            json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8"
        )

    print(f"Hasil ekstraksi ditulis ke {OUT_DIR}")
    for filename, data in payload.items():
        print(f"  {filename:28} {len(data):>5} baris")

    total_payroll = sum(record["wageAmount"] for record in payroll)
    total_expenses = sum(txn["amount"] for txn in expenses)
    total_capital = sum(entry["amount"] for entry in capital)
    total_revenue = sum(item["totalRevenue"] for item in payload["harvests.json"])
    print()
    print(f"  Total dana masuk       Rp {total_capital:>15,}")
    print(f"  Total pengeluaran kas  Rp {total_expenses:>15,}")
    print(f"  Total upah harian      Rp {total_payroll:>15,}")
    print(f"  Total pendapatan panen Rp {total_revenue:>15,}")
    print(f"  Peringatan             {len(warnings)}")


if __name__ == "__main__":
    main()
